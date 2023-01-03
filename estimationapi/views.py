from posixpath import split
from django.http import HttpResponse
from django.shortcuts import render
import openpyxl
from estimationapi.serializers import PopulationActiveOffreEmploiSecteurSerializer
from estimationapi.models import DonneeConsommation, DonneeInvestissementPrive,ImpotDGI
from estimationapi.models import DonneeExportation, DonneeImportation,RevenusSalaires,ExedantBrutExploitation,AutresImpots,SubventionProduction,SubventionConsommation
from estimationapi.serializers import DonneeImportationSerializer
from estimationapi.serializers import DonneeExportationSerializer
from estimationapi.models import DonneeDepenseCapitale, DonneeDepenseCourante
from estimationapi.serializers import DonneeDepenseCapitaleSerializer, DonneeDepenseCouranteSerializer
from estimationapi.serializers import DonneeInvestissementPriveSerializer
from estimationapi.serializers import DonneeConsommationSerializer
from estimationapi.serializers import PopulationActiveOffreEmploiSerializer
from estimationapi.serializers import TauxMigrationProvinceSerializer
from estimationapi.models import TauxMigrationProvince
from estimationapi.serializers import TauxCroissanceProvinceSerializer
from estimationapi.models import PopulationProvince
from estimationapi.models import TauxNataliteMortalite
from estimationapi.serializers import TauxNataliteMortaliteSerializer
from estimationapi.serializers import PopulationProvinceSerializer,RevenusSalairesSerializer,ExedantBrutExploitationSerializer,AutresImpotsSerializer,SubventionProductionSerializer,SubventionConsommationSerializer
from estimationapi.serializers import ImpotDGISerializer,ImpoExpoDGDASerializer,RecettesDGDASerializer,RecettesDGRADSerializer
from rest_framework.response import Response
from rest_framework.decorators import api_view
from knox.views import LoginView as KnoxLoginView
from rest_framework.authtoken.serializers import AuthTokenSerializer
from rest_framework import generics, permissions
from .serializers import  DonneeProductionSerializer, UserSerializer, RegisterSerializer
from django.contrib.auth import login
from knox.models import AuthToken
from rest_framework.authtoken.serializers import AuthTokenSerializer

# Create your views here.


class LoginAPI(KnoxLoginView):
    permission_classes = (permissions.AllowAny,)

    def post(self, request, format=None):
        serializer = AuthTokenSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']
        login(request, user)
        return Response({"username" : user.username})

# Register API
class RegisterAPI(generics.GenericAPIView):
    serializer_class = RegisterSerializer

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response({
        "user": UserSerializer(user, context=self.get_serializer_context()).data,
        "token": AuthToken.objects.create(user)[1]
        })

def welcom(request):
    return HttpResponse('WELCOM TO MODELISATION  API RESOURCES')        


@api_view(['POST'])   
def population(request): 
    fichier = request.data['fichier']
    serializer = PopulationProvinceSerializer(data={'fichier':fichier})
    if serializer.is_valid() :
        serializer.save()
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(2,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"taille_province":sheet.cell(i,2).value,"split_urbain":sheet.cell(i,3).value,"split_rural":sheet.cell(i,4).value});          
    return Response(donnes)

@api_view(['POST'])   
def tauxNataliteMortalite(request): 
    fichier = request.data['fichier']
    serializer = TauxNataliteMortaliteSerializer(data={'fichier':fichier})
    if serializer.is_valid() :
        serializer.save()
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(2,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"taux_natalite_urbain":sheet.cell(i,2).value,"taux_mortalite_urbain":sheet.cell(i,3).value,"taux_natalite_rural":sheet.cell(i,4).value,"taux_mortalite_rural":sheet.cell(i,5).value});          
    return Response(donnes)


@api_view(['POST'])   
def tauxMigrationProvince(request): 
    fichier = request.data['fichier']
    serializer = TauxMigrationProvinceSerializer(data={'fichier':fichier})
    if serializer.is_valid() :
        serializer.save()
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(2,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"taux_migration":sheet.cell(i,2).value});          
    return Response(donnes)

@api_view(['POST'])   
def populationActiveOffre(request): 
    fichier = request.data['fichier']
    serializer = PopulationActiveOffreEmploiSerializer(data={'fichier':fichier})
    if serializer.is_valid() :
        serializer.save()
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(2,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"population_active_urbaine_participation":sheet.cell(i,2).value,"population_active_rurale_participation":sheet.cell(i,3).value});          
    return Response(donnes)

@api_view(['POST'])   
def populationActiveOffreSecteur(request): 
    fichier = request.data['fichier']
    serializer = PopulationActiveOffreEmploiSecteurSerializer(data={'fichier':fichier})
    if serializer.is_valid() :
        serializer.save()
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(3,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"population_active_urbaine_participation_agriculture":sheet.cell(i,2).value,"population_active_urbaine_participation_mines":sheet.cell(i,3).value,"population_active_urbaine_participation_industrie":sheet.cell(i,4).value,"population_active_urbaine_participation_service":sheet.cell(i,5).value,"population_active_rurale_participation_agriculture":sheet.cell(i,6).value,"population_active_rurale_participation_mines":sheet.cell(i,7).value,"population_active_rurale_participation_industrie":sheet.cell(i,8).value,"population_active_rurale_participation_service":sheet.cell(i,9).value});          
    return Response(donnes)


@api_view(['POST'])   
def donneeConsommation(request): 
    fichier = request.data['fichier']
    serializer = DonneeConsommationSerializer(data={'fichier':fichier})
    if serializer.is_valid() :
        serializer.save()
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(3,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"nature_consommation":sheet.cell(i,2).value,"milieu_urbaine_alimentaire":sheet.cell(i,3).value,"milieu_urbaine_biens_durables":sheet.cell(i,4).value,"milieu_urbaine_autres_biens":sheet.cell(i,5).value,"milieu_rurale_alimentaire":sheet.cell(i,6).value,"milieu_rurale_biens_durables":sheet.cell(i,7).value,"milieu_rurale_autres_biens":sheet.cell(i,8).value});          
    return Response(donnes)

@api_view(['POST'])   
def donneeInvestissementPrive(request): 
    fichier = request.data['fichier']
    serializer = DonneeInvestissementPriveSerializer(data={'fichier':fichier})
    if serializer.is_valid() :
        serializer.save()
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(3,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"nature_investissement_prive":sheet.cell(i,2).value,"milieu_urbaine_agriculture":sheet.cell(i,3).value,"milieu_urbaine_mines":sheet.cell(i,4).value,"milieu_urbaine_industrie":sheet.cell(i,5).value,"milieu_urbaine_service":sheet.cell(i,6).value,"milieu_rurale_agriculture":sheet.cell(i,7).value,"milieu_rurale_mines":sheet.cell(i,8).value,"milieu_rurale_industrie":sheet.cell(i,9).value,"milieu_rurale_service":sheet.cell(i,10).value});
    return Response(donnes)

@api_view(['POST'])   
def donneeDepenseCourante(request): 
    fichier = request.data['fichier']
    serializer = DonneeDepenseCouranteSerializer(data={'fichier':fichier})
    if serializer.is_valid() :
        serializer.save()
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(3,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"nature_depense_courante":sheet.cell(i,2).value,"milieu_urbaine_agriculture":sheet.cell(i,3).value,"milieu_urbaine_mines":sheet.cell(i,4).value,"milieu_urbaine_industrie":sheet.cell(i,5).value,"milieu_urbaine_service":sheet.cell(i,6).value,"milieu_rurale_agriculture":sheet.cell(i,7).value,"milieu_rurale_mines":sheet.cell(i,8).value,"milieu_rurale_industrie":sheet.cell(i,9).value,"milieu_rurale_service":sheet.cell(i,10).value});          
    return Response(donnes)

@api_view(['POST'])   
def donneeDepenseCapital(request): 
    fichier = request.data['fichier']
    serializer = DonneeDepenseCapitaleSerializer(data={'fichier':fichier})
    if serializer.is_valid() :
        serializer.save()
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(3,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"nature_depense_capital":sheet.cell(i,2).value,"milieu_urbaine_agriculture":sheet.cell(i,3).value,"milieu_urbaine_mines":sheet.cell(i,4).value,"milieu_urbaine_industrie":sheet.cell(i,5).value,"milieu_urbaine_service":sheet.cell(i,6).value,"milieu_rurale_agriculture":sheet.cell(i,7).value,"milieu_rurale_mines":sheet.cell(i,8).value,"milieu_rurale_industrie":sheet.cell(i,9).value,"milieu_rurale_service":sheet.cell(i,10).value});
    return Response(donnes)

@api_view(['POST'])   
def donneeExportation(request): 
    fichier = request.data['fichier']
    serializer = DonneeExportationSerializer(data={'fichier':fichier})
    if serializer.is_valid() :
        serializer.save()
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(3,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"nature_exportations":sheet.cell(i,2).value,"milieu_urbaine_agriculture":sheet.cell(i,3).value,"milieu_urbaine_mines":sheet.cell(i,4).value,"milieu_urbaine_industrie":sheet.cell(i,5).value,"milieu_urbaine_service":sheet.cell(i,6).value,"milieu_rurale_agriculture":sheet.cell(i,7).value,"milieu_rurale_mines":sheet.cell(i,8).value,"milieu_rurale_industrie":sheet.cell(i,9).value,"milieu_rurale_service":sheet.cell(i,10).value});          
    return Response(donnes)

@api_view(['POST'])   
def donneeImportation(request): 
    fichier = request.data['fichier']
    serializer = DonneeImportationSerializer(data={'fichier':fichier})
    if serializer.is_valid() :
        serializer.save()
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(3,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"nature_importations":sheet.cell(i,2).value,"milieu_urbaine_agriculture":sheet.cell(i,3).value,"milieu_urbaine_mines":sheet.cell(i,4).value,"milieu_urbaine_industrie":sheet.cell(i,5).value,"milieu_urbaine_service":sheet.cell(i,6).value,"milieu_rurale_agriculture":sheet.cell(i,7).value,"milieu_rurale_mines":sheet.cell(i,8).value,"milieu_rurale_industrie":sheet.cell(i,9).value,"milieu_rurale_service":sheet.cell(i,10).value});          
    return Response(donnes)


@api_view(['POST'])   
def donneeProduction(request): 
    fichier = request.data['fichier']
    serializer = DonneeProductionSerializer(data={'fichier':fichier})
    if serializer.is_valid() :
        serializer.save()
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(3,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"nature_production":sheet.cell(i,2).value,"milieu_urbaine_agriculture":sheet.cell(i,3).value,"milieu_urbaine_mines":sheet.cell(i,4).value,"milieu_urbaine_industrie":sheet.cell(i,5).value,"milieu_urbaine_service":sheet.cell(i,6).value,"milieu_rurale_agriculture":sheet.cell(i,7).value,"milieu_rurale_mines":sheet.cell(i,8).value,"milieu_rurale_industrie":sheet.cell(i,9).value,"milieu_rurale_service":sheet.cell(i,10).value});          
    return Response(donnes)




@api_view(['GET'])   
def tauxCroissance(request): 
    tauxNataliteMortalite = TauxNataliteMortalite.objects.all()
    fichier = ""
    for x in tauxNataliteMortalite:
            fichier = x.fichier
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(2,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"taux_natalite_urbain":sheet.cell(i,2).value,"taux_mortalite_urbain":sheet.cell(i,3).value,"taux_natalite_rural":sheet.cell(i,4).value,"taux_mortalite_rural":sheet.cell(i,5).value});          
    return Response(donnes)

@api_view(['GET'])   
def tailleReellePopulation(request): 
    populationProvince = PopulationProvince.objects.all()
    fichier = ""
    for x in populationProvince:
            fichier = x.fichier
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(2,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"taille_province":sheet.cell(i,2).value,"split_urbain":sheet.cell(i,3).value,"split_rural":sheet.cell(i,4).value});          
    return Response(donnes)

@api_view(['GET'])   
def tauxMigration(request): 
    tauxMigrationProvince = TauxMigrationProvince.objects.all()
    fichier = ""
    for x in tauxMigrationProvince:
            fichier = x.fichier
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(2,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"taux_migration":sheet.cell(i,2).value});          
    return Response(donnes)

@api_view(['GET'])   
def depenseCourante(request): 
    donneeDepenseCourante = DonneeDepenseCourante.objects.all()
    fichier = ""
    for x in donneeDepenseCourante:
            fichier = x.fichier
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(3,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"nature_depense_courante":sheet.cell(i,2).value,"milieu_urbaine_agriculture":sheet.cell(i,3).value,"milieu_urbaine_mines":sheet.cell(i,4).value,"milieu_urbaine_industrie":sheet.cell(i,5).value,"milieu_urbaine_service":sheet.cell(i,6).value,"milieu_rurale_agriculture":sheet.cell(i,8).value,"milieu_rurale_mines":sheet.cell(i,9).value,"milieu_rurale_industrie":sheet.cell(i,10).value,"milieu_rurale_service":sheet.cell(i,11).value});          
    return Response(donnes)

@api_view(['GET'])   
def depenseCapital(request): 
    donneeDepenseCapital = DonneeDepenseCapitale.objects.all()
    fichier = ""
    for x in donneeDepenseCapital:
            fichier = x.fichier
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(3,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"nature_depense_capital":sheet.cell(i,2).value,"milieu_urbaine_agriculture":sheet.cell(i,3).value,"milieu_urbaine_mines":sheet.cell(i,4).value,"milieu_urbaine_industrie":sheet.cell(i,5).value,"milieu_urbaine_service":sheet.cell(i,6).value,"milieu_rurale_agriculture":sheet.cell(i,8).value,"milieu_rurale_mines":sheet.cell(i,9).value,"milieu_rurale_industrie":sheet.cell(i,10).value,"milieu_rurale_service":sheet.cell(i,11).value});          
    return Response(donnes)

@api_view(['GET'])   
def donneeExportationAuto(request): 
    donneeExportation = DonneeExportation.objects.all()
    fichier = ""
    for x in donneeExportation:
            fichier = x.fichier
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(3,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"nature_exportations":sheet.cell(i,2).value,"milieu_urbaine_agriculture":sheet.cell(i,3).value,"milieu_urbaine_mines":sheet.cell(i,4).value,"milieu_urbaine_industrie":sheet.cell(i,5).value,"milieu_urbaine_service":sheet.cell(i,6).value,"milieu_rurale_agriculture":sheet.cell(i,8).value,"milieu_rurale_mines":sheet.cell(i,9).value,"milieu_rurale_industrie":sheet.cell(i,10).value,"milieu_rurale_service":sheet.cell(i,11).value});          
    return Response(donnes)

@api_view(['GET'])   
def donneeImportationAuto(request): 
    donneeImportation = DonneeImportation.objects.all()
    fichier = ""
    for x in donneeImportation:
            fichier = x.fichier
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(3,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"nature_importations":sheet.cell(i,2).value,"milieu_urbaine_agriculture":sheet.cell(i,3).value,"milieu_urbaine_mines":sheet.cell(i,4).value,"milieu_urbaine_industrie":sheet.cell(i,5).value,"milieu_urbaine_service":sheet.cell(i,6).value,"milieu_rurale_agriculture":sheet.cell(i,8).value,"milieu_rurale_mines":sheet.cell(i,9).value,"milieu_rurale_industrie":sheet.cell(i,10).value,"milieu_rurale_service":sheet.cell(i,11).value});          
    return Response(donnes)


@api_view(['GET'])   
def donneeConsommationAuto(request): 
    donneeConsommation = DonneeConsommation.objects.all()
    fichier = ""
    for x in donneeConsommation:
            fichier = x.fichier
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(3,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"nature_consommation":sheet.cell(i,2).value,"milieu_urbaine_agriculture":sheet.cell(i,3).value,"milieu_urbaine_mines":sheet.cell(i,4).value,"milieu_urbaine_industrie":sheet.cell(i,5).value,"milieu_urbaine_service":sheet.cell(i,6).value,"milieu_rurale_agriculture":sheet.cell(i,8).value,"milieu_rurale_mines":sheet.cell(i,9).value,"milieu_rurale_industrie":sheet.cell(i,10).value,"milieu_rurale_service":sheet.cell(i,11).value});          
    return Response(donnes)

@api_view(['GET'])   
def donneeInvestissementPriveAuto(request): 
    donneeInvestissementPrive = DonneeInvestissementPrive.objects.all()
    fichier = ""
    for x in donneeInvestissementPrive:
            fichier = x.fichier
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(3,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"nature_investissement_prive":sheet.cell(i,2).value,"milieu_urbaine_agriculture":sheet.cell(i,3).value,"milieu_urbaine_mines":sheet.cell(i,4).value,"milieu_urbaine_industrie":sheet.cell(i,5).value,"milieu_urbaine_service":sheet.cell(i,6).value,"milieu_rurale_agriculture":sheet.cell(i,8).value,"milieu_rurale_mines":sheet.cell(i,9).value,"milieu_rurale_industrie":sheet.cell(i,10).value,"milieu_rurale_service":sheet.cell(i,11).value});          
    return Response(donnes)


@api_view(['POST'])   
def impotDGI(request): 
    fichier = request.data['fichier']
    serializer = ImpotDGISerializer(data={'fichier':fichier})
    if serializer.is_valid() :
        serializer.save()
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(2,row + 1):
            donnes.append({"recettes_dgi":sheet.cell(i,1).value,"annee_fiscale_2018":sheet.cell(i,2).value,"annee_fiscale_2019":sheet.cell(i,3).value,"annee_fiscale_2020":sheet.cell(i,4).value,"annee_fiscale_2021":sheet.cell(i,5).value});          
    return Response(donnes)

@api_view(['POST'])   
def impoExpoDGDA(request): 
    fichier = request.data['fichier']
    serializer = ImpoExpoDGDASerializer(data={'fichier':fichier})
    if serializer.is_valid() :
        serializer.save()
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(2,row + 1):
            donnes.append({"expo_impo_dgda":sheet.cell(i,1).value,"annee_fiscale_2018":sheet.cell(i,2).value,"annee_fiscale_2019":sheet.cell(i,3).value,"annee_fiscale_2020":sheet.cell(i,4).value,"annee_fiscale_2021":sheet.cell(i,5).value});          
    return Response(donnes)

@api_view(['POST'])   
def recettesDGDA(request): 
    fichier = request.data['fichier']
    serializer = RecettesDGDASerializer(data={'fichier':fichier})
    if serializer.is_valid() :
        serializer.save()
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(2,row + 1):
            donnes.append({"recettes_dgda":sheet.cell(i,1).value,"annee_fiscale_2018":sheet.cell(i,2).value,"annee_fiscale_2019":sheet.cell(i,3).value,"annee_fiscale_2020":sheet.cell(i,4).value,"annee_fiscale_2021":sheet.cell(i,5).value});          
    return Response(donnes)

@api_view(['POST'])   
def recettesDGRAD(request): 
    fichier = request.data['fichier']
    serializer = RecettesDGRADSerializer(data={'fichier':fichier})
    if serializer.is_valid() :
        serializer.save()
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(2,row + 1):
            donnes.append({"recettes_dgrad":sheet.cell(i,1).value,"annee_fiscale_2019":sheet.cell(i,2).value,"annee_fiscale_2020":sheet.cell(i,3).value,"annee_fiscale_2021":sheet.cell(i,4).value,"annee_fiscale_2022":sheet.cell(i,5).value});          
    return Response(donnes)

@api_view(['POST'])   
def revenusSalaires(request): 
    fichier = request.data['fichier']
    serializer = RevenusSalairesSerializer(data={'fichier':fichier})
    if serializer.is_valid() :
        serializer.save()
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(3,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"nature_revenus_salaires":sheet.cell(i,2).value,"salaires_urbain_agricultures":sheet.cell(i,3).value,"salaire_urbain_insdustries_extractives":sheet.cell(i,4).value,"salaires_urbain_industries_manufactures":sheet.cell(i,5).value,"salaires_urbain_services":sheet.cell(i,6).value,"salaires_rural_agricultures":sheet.cell(i,7).value,"salaire_rural_insdustries_extractives":sheet.cell(i,8).value,"salaires_rural_industries_manufactures":sheet.cell(i,9).value,"salaires_rural_services":sheet.cell(i,10).value});          
    return Response(donnes)

@api_view(['POST'])   
def exedantBruteExploitation(request): 
    fichier = request.data['fichier']
    serializer = RevenusSalairesSerializer(data={'fichier':fichier})
    if serializer.is_valid() :
        serializer.save()
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(3,row + 1):
        donnes.append({"nom_province":sheet.cell(i,1).value,"nature_revenus_salaires":sheet.cell(i,2).value,"salaires_urbain_agricultures":sheet.cell(i,3).value,"salaire_urbain_insdustries_extractives":sheet.cell(i,4).value,"salaires_urbain_industries_manufactures":sheet.cell(i,5).value,"salaires_urbain_services":sheet.cell(i,6).value,"salaires_rural_agricultures":sheet.cell(i,7).value,"salaire_rural_insdustries_extractives":sheet.cell(i,8).value,"salaires_rural_industries_manufactures":sheet.cell(i,9).value,"salaires_rural_services":sheet.cell(i,10).value});          
           # donnes.append({"nom_province":sheet.cell(i,1).value,"nature_exedant_brut":sheet.cell(i,2).value,"ebe_urbain_agricultures":sheet.cell(i,3).value,"ebe_urbain_insdustries_extractives":sheet.cell(i,4).value,"ebe_urbain_industries_manufactures":sheet.cell(i,5).value,"ebe_urbain_services":sheet.cell(i,6).value,"ebe_rural_agricultures":sheet.cell(i,7).value,"ebe_rural_insdustries_extractives":sheet.cell(i,8).value,"ebe_rural_industries_manufactures":sheet.cell(i,9).value,"ebe_rural_services":sheet.cell(i,10).value});          
    return Response(donnes)

@api_view(['POST'])   
def autresImpots(request): 
    fichier = request.data['fichier']
    serializer = RevenusSalairesSerializer(data={'fichier':fichier})
    if serializer.is_valid() :
        serializer.save()
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(3,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"nature_autre_impot":sheet.cell(i,2).value,"autre_impot_urbain":sheet.cell(i,3).value,"autre_impot_urbain_total":sheet.cell(i,4).value,"autre_impot_rural":sheet.cell(i,5).value,"autre_impot_rural_total":sheet.cell(i,6).value});          
    return Response(donnes)

@api_view(['POST'])   
def subventionProduction(request): 
    fichier = request.data['fichier']
    serializer = RevenusSalairesSerializer(data={'fichier':fichier})
    if serializer.is_valid() :
        serializer.save()
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(3,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"nature_subvention_production":sheet.cell(i,2).value,"ebe_urbain_agricultures":sheet.cell(i,3).value,"ebe_urbain_insdustries_extractives":sheet.cell(i,4).value,"ebe_urbain_industries_manufactures":sheet.cell(i,5).value,"ebe_urbain_services":sheet.cell(i,6).value,"ebe_rural_agricultures":sheet.cell(i,7).value,"ebe_rural_insdustries_extractives":sheet.cell(i,8).value,"ebe_rural_industries_manufactures":sheet.cell(i,9).value,"ebe_rural_services":sheet.cell(i,10).value});          
    return Response(donnes)

@api_view(['POST'])   
def subventionConsommation(request): 
    fichier = request.data['fichier']
    serializer = RevenusSalairesSerializer(data={'fichier':fichier})
    if serializer.is_valid() :
        serializer.save()
    wb= openpyxl.load_workbook(fichier)
    sheet = wb['Feuil1']
    row = sheet.max_row
    donnes = []
    for i in range(2,row + 1):
            donnes.append({"nom_province":sheet.cell(i,1).value,"nature_subvention_consommation":sheet.cell(i,2).value,"subvention_consommation_urbain":sheet.cell(i,3).value,"subvention_consommation_rural":sheet.cell(i,4).value});          
    return Response(donnes)






